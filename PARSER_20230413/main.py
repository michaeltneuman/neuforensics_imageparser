import argparse
import csv
import datetime
import logging
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import os
import requests
import shutil
import subprocess
import time
import win32api
import zipfile


def get_current_time_as_string():
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S")


def logging_configuration(report_start_time_lvariable):
    logfilename = "MTN_IMAGE_PARSE_" + report_start_time_lvariable + '.log'
    logging.basicConfig(handlers=[logging.FileHandler(logfilename), logging.StreamHandler()], level=logging.INFO,
                        format='%(asctime)s | %(levelname)s | %(message)s')
    logging.info(f"Log file {logfilename} created.")


def argument_handler():
    csv.field_size_limit(900000000)
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--image", help="Image File (E01)", required=True)
    parser.add_argument("-t", "--tools", help="Tools Directory (default = current directory)",
                        default=os.getcwd()+"\\tools")
    parser.add_argument("-o", "--output", help="Output Directory (default = current directory)",
                        default=os.getcwd()+"\\output")
    args = parser.parse_args()
    if not os.path.exists(args.tools):
        os.mkdir(args.tools)
    if not os.path.exists(args.output):
        os.mkdir(args.output)
    logging.info("--image: " + os.path.abspath(args.image) + "\t--tools: " + os.path.abspath(args.tools) + '\\'
                 + "\t--output: " + os.path.abspath(args.output)+'\\')
    return os.path.abspath(args.image), os.path.abspath(args.tools)+'\\', os.path.abspath(args.output)+'\\'


def download_url(url, output_directory_local):
    return_filepath = output_directory_local+'temp.' + url.split('.')[-1]
    logging.info("Downloading " + url + " to " + return_filepath)
    with open(return_filepath, 'wb') as outfile:
        shutil.copyfileobj(requests.get(url, stream=True, timeout=5).raw, outfile)
    logging.info("Download complete.")
    return return_filepath


def unzip_file(zip_file, extract_location="", delete_original=False):
    logging.info("Unzipping " + zip_file + " to " + extract_location + "...")
    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
        zip_ref.extractall(extract_location)
    logging.info("Unzip complete.")
    if delete_original:
        logging.info("Deleting original file " + zip_file + "...")
        os.remove(zip_file)
        logging.info("Original file deleted.")
    return extract_location


def get_list_of_drive_letters():
    drive_letter_list = set(win32api.GetLogicalDriveStrings().split('\000')[:-1])
    logging.info("Drive letters found: " + str(drive_letter_list))
    return drive_letter_list


def mount_image_file_locally(image_file_lvariable, tools_dir_lvariable):
    aim_cli_path = tools_dir_lvariable+"Arsenal-Image-Mounter-v3.9.239\\aim_cli.exe"
    while not os.path.exists(aim_cli_path):
        try:
            logging.info("Downloading Arsenal Image Mounter...")
            subprocess.run([os.environ['LOCALAPPDATA']+"\\MEGAcmd\\MEGAclient.exe", "get",
                            "https://mega.nz/file/T1gBDIhZ#lbOrVwmPx8OF0rwCEFieNFnGiwycdcahwFVbvtiRI_A",
                            tools_dir_lvariable])
            logging.info("Arsenal Image Mounter downloaded.")
            unzip_file(os.path.join(tools_dir_lvariable, "Arsenal-Image-Mounter-v3.9.239.zip"), tools_dir_lvariable,
                       delete_original=True)
            break
        except Exception as e:
            logging.warning(f"Error {str(e)}, trying to download MegaCMD...")
            installer_filepath = download_url("https://mega.nz/MEGAcmdSetup64.exe", tools_dir_lvariable)
            logging.info("MEGAcmd downloaded, installing...")
            subprocess.run([installer_filepath, "/S"])
            logging.info("MEGAcmd installed.")
    subprocess.run([aim_cli_path, "--dismount"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    time.sleep(5)
    drive_letter, starting_drive_letters = None, get_list_of_drive_letters().copy()
    os.chdir(os.path.dirname(image_file_lvariable))
    while drive_letter is None:
        subprocess.Popen([aim_cli_path, "--mount", "--filename=" + os.path.basename(image_file_lvariable)],
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(5)
        for new_drive in get_list_of_drive_letters():
            if new_drive not in starting_drive_letters:
                drive_letter = new_drive
                logging.info("Image mounted to " + drive_letter)
                break
    return drive_letter
    # figure out how to uninstall megacmd


def net6_installed():
    return subprocess.check_output(["dotnet", "--list-runtimes"]).decode("utf-8").find("Microsoft.NETCore.App 6.") != -1


def install_net6(tools_dir_lvariable):
    download_url("""https://download.visualstudio.microsoft.com/download/pr/85473c45-8d91-48cb-ab41-86ec7abc1000/
                 83cd0c82f0cde9a566bae4245ea5a65b/windowsdesktop-runtime-6.0.16-win-x64.exe""", tools_dir_lvariable)
    logging.info("Installing .NET6 Runtime...")
    subprocess.run([tools_dir_lvariable + "windowsdesktop-runtime-6.0.16-win-x64.exe", "/quiet", "/norestart"])
    logging.info(".NET6 Runtime installed.")


# need to modify to go by md5
def tool_installed(tool_lvariable, tools_dir_lvariable):
    for dirpath, dirnames, filenames in os.walk(tools_dir_lvariable):
        for filename in filenames:
            if tool_lvariable == filename:
                return True
    return False


def determine_file_to_open(report_title):
    if report_title in ('Amcache'):
        return 'temp_UnassociatedFileEntries.csv'
    else:
        return 'temp.csv'


def remove_bad_characters(list_of_entries):
    return list_of_entries


def jumplist_report_handler(output_directory_local, workbook_local,report_title):
    max_rows_per_sheet, chunksize = 500000, 50000
    for file in os.listdir(output_directory_local):
        chunk,num_rows,additional_sheet_num = [],0,2
        if '_AutomaticDestinations' in file or 'temp_CustomDestinations' in file:
            with open(output_directory_local+file,'r',encoding='utf-8-sig') as f:
                sheet = workbook_local.create_sheet(title=f'{report_title}{file.split("_")[-1]}')
                reader = csv.reader(f)
                headers = next(reader)
                sheet.append(headers)
                logging.info("Headers = " + str(headers))
                for row in reader:
                    chunk.append(row)
                    if len(chunk) % chunksize == 0:
                        logging.info(f"{str(num_rows + chunksize)} rows chunked.")
                        for row in chunk:
                            try:
                                sheet.append(row)
                                num_rows += 1
                            except IllegalCharacterError as e:
                                illegal_char, revised_row = str(e).split(" cannot be used in worksheets.")[0], []
                                for element in row:
                                    if illegal_char in element:
                                        revised_row.append(element.replace(illegal_char, '!'))
                                        logging.warning(f"Removed bad character from row {str(num_rows)}")
                                    else:
                                        revised_row.append(element)
                                sheet.append(revised_row)
                                num_rows += 1
                        if num_rows % max_rows_per_sheet == 0:
                            logging.info(
                                f"Max rows reached! Creating new sheet {report_title}{str(additional_sheet_num)}")
                            sheet = workbook_local.create_sheet(title=f'{report_title}{file.split("_")[-1]}{str(additional_sheet_num)}')
                            sheet.append(headers)
                            additional_sheet_num += 1
                        del chunk
                        chunk = []
                        logging.info("Saving workbook...")
                        workbook.save(output_directory_local + 'output.xlsx')
                        logging.info("Workbook saved to " + output_directory_local + 'output.xlsx')
                for row in chunk:
                    try:
                        sheet.append(row)
                    except IllegalCharacterError as e:
                        illegal_char, revised_row = e.args[0].split(" cannot be used in worksheets.")[0], []
                        for element in row:
                            if illegal_char in element:
                                revised_row.append(element.replace(illegal_char, '!'))
                            else:
                                revised_row.append(element)
                        try:
                            sheet.append(revised_row)
                            logging.warning(f"Removed bad character from row {str(num_rows)}")
                        except IllegalCharacterError as e:
                            logging.error(f"Could not append row {str(num_rows)} to sheet due to illegal character.")
                    num_rows += 1
                del chunk
            logging.info(f'{report_title}{file.split("_")[-1]} FINISHED! {str(num_rows)} total rows written to workbook.')
    return workbook_local

def srum_report_handler(output_directory_local, workbook_local,report_title):
    max_rows_per_sheet, chunksize = 500000, 50000
    for file in os.listdir(output_directory_local):
        chunk,num_rows,additional_sheet_num = [],0,2
        if '_SrumECmd_' in file:
            with open(output_directory_local+file,'r',encoding='utf-8-sig') as f:
                sheet = workbook_local.create_sheet(title=f'{report_title}{file.split("_SrumECmd_")[-1].replace("_Output.csv","")}')
                reader = csv.reader(f)
                headers = next(reader)
                sheet.append(headers)
                logging.info("Headers = " + str(headers))
                for row in reader:
                    chunk.append(row)
                    if len(chunk) % chunksize == 0:
                        logging.info(f"{str(num_rows + chunksize)} rows chunked.")
                        for row in chunk:
                            try:
                                sheet.append(row)
                                num_rows += 1
                            except IllegalCharacterError as e:
                                illegal_char, revised_row = str(e).split(" cannot be used in worksheets.")[0], []
                                for element in row:
                                    if illegal_char in element:
                                        revised_row.append(element.replace(illegal_char, '!'))
                                        logging.warning(f"Removed bad character from row {str(num_rows)}")
                                    else:
                                        revised_row.append(element)
                                sheet.append(revised_row)
                                num_rows += 1
                        if num_rows % max_rows_per_sheet == 0:
                            logging.info(
                                f"Max rows reached! Creating new sheet {report_title}{str(additional_sheet_num)}")
                            sheet = workbook_local.create_sheet(title=f'{report_title}{file.split("_SrumECmd_")[-1].replace("_Output.csv","")}{str(additional_sheet_num)}')
                            sheet.append(headers)
                            additional_sheet_num += 1
                        del chunk
                        chunk = []
                        logging.info("Saving workbook...")
                        workbook.save(output_directory_local + 'output.xlsx')
                        logging.info("Workbook saved to " + output_directory_local + 'output.xlsx')
                for row in chunk:
                    try:
                        sheet.append(row)
                    except IllegalCharacterError as e:
                        illegal_char, revised_row = e.args[0].split(" cannot be used in worksheets.")[0], []
                        for element in row:
                            if illegal_char in element:
                                revised_row.append(element.replace(illegal_char, '!'))
                            else:
                                revised_row.append(element)
                        try:
                            sheet.append(revised_row)
                            logging.warning(f"Removed bad character from row {str(num_rows)}")
                        except IllegalCharacterError as e:
                            logging.error(f"Could not append row {str(num_rows)} to sheet due to illegal character.")
                    num_rows += 1
                del chunk
            logging.info(f'{report_title}{file.split("_SrumECmd_")[-1].replace("_Output.csv","")} FINISHED! {str(num_rows)} total rows written to workbook.')
    return workbook_local

def shellbags_report_handler(output_directory_local, workbook_local,report_title):
    max_rows_per_sheet, chunksize = 500000, 50000
    sheet = workbook_local.create_sheet(title=f'{report_title}')
    headers = ['BagPath',	'Slot',	'NodeSlot',	'MRUPosition',	'AbsolutePath',	'ShellType',	'Value',	'ChildBags',	'CreatedOn',	'ModifiedOn',	'AccessedOn',	'LastWriteTime',	'MFTEntry',	'MFTSequenceNumber',	'ExtensionBlockCount',	'FirstInteracted',	'LastInteracted',	'HasExplored',	'Miscellaneous',	'SourceFile']
    sheet.append(headers)
    for file in os.listdir(output_directory_local):
        chunk,num_rows,additional_sheet_num = [],0,2
        if '_NTUSER_temp' in file or '_UsrClass_temp' in file:
            fname = output_directory_local+file
            with open(fname,'r',encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                next(reader)
                for row in reader:
                    chunk.append(row+[fname])
                    if len(chunk) % chunksize == 0:
                        logging.info(f"{str(num_rows + chunksize)} rows chunked.")
                        for row in chunk:
                            try:
                                sheet.append(row)
                                num_rows += 1
                            except IllegalCharacterError as e:
                                illegal_char, revised_row = str(e).split(" cannot be used in worksheets.")[0], []
                                for element in row:
                                    if illegal_char in element:
                                        revised_row.append(element.replace(illegal_char, '!'))
                                        logging.warning(f"Removed bad character from row {str(num_rows)}")
                                    else:
                                        revised_row.append(element)
                                sheet.append(revised_row)
                                num_rows += 1
                        if num_rows % max_rows_per_sheet == 0:
                            logging.info(
                                f"Max rows reached! Creating new sheet {report_title}{str(additional_sheet_num)}")
                            sheet = workbook_local.create_sheet(title=f'{report_title}{str(additional_sheet_num)}')
                            sheet.append(headers)
                            additional_sheet_num += 1
                        del chunk
                        chunk = []
                        logging.info("Saving workbook...")
                        workbook.save(output_directory_local + 'output.xlsx')
                        logging.info("Workbook saved to " + output_directory_local + 'output.xlsx')
                for row in chunk:
                    try:
                        sheet.append(row)
                    except IllegalCharacterError as e:
                        illegal_char, revised_row = e.args[0].split(" cannot be used in worksheets.")[0], []
                        for element in row:
                            if illegal_char in element:
                                revised_row.append(element.replace(illegal_char, '!'))
                            else:
                                revised_row.append(element)
                        try:
                            sheet.append(revised_row)
                            logging.warning(f"Removed bad character from row {str(num_rows)}")
                        except IllegalCharacterError as e:
                            logging.error(f"Could not append row {str(num_rows)} to sheet due to illegal character.")
                    num_rows += 1
                del chunk
            logging.info(f'{report_title} FINISHED! {str(num_rows)} total rows written to workbook.')
    return workbook_local

def execute_tool(tool_exe_filepath, relevant_artifact_path, output_dir_lvariable, workbook_local, report_title):
    logging.info("Executing " + tool_exe_filepath + "...")
    enc = 'utf-8-sig'
    flag1 = '-f'
    if isinstance(relevant_artifact_path, list):
        flag1,relevant_artifact_path = '-d',relevant_artifact_path[0]
    command_to_run = [tool_exe_filepath, flag1, relevant_artifact_path, '--csv', output_dir_lvariable,
                      '--csvf', 'temp.csv']
    if tool_exe_filepath.endswith("BrowsingHistoryView.exe"):
        enc = 'latin-1'
        command_to_run = [tool_exe_filepath,'/HistorySource','2',relevant_artifact_path,'/scomma',output_dir_lvariable+'temp.csv']
    if tool_exe_filepath.endswith("SrumECmd.exe"):
        command_to_run = [tool_exe_filepath, '-f', relevant_artifact_path,'-r',relevant_artifact_path.split('\\')[0]+
                          "\\windows\\system32\\config\\software",'--csv',output_dir_lvariable]
    logging.info(command_to_run)
    result = subprocess.run(command_to_run, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if "Found 0 files" in result.stdout.decode('utf-8'):
        logging.warning("No files found. Skipping...")
    else:
        logging.info(
            "Execution complete. Parsing output to Excel file...Creating new sheet...Opening csv file and reading rows...")
        if report_title in ("SRUM"):
            workbook_local = srum_report_handler(output_dir_lvariable, workbook_local,report_title)
        elif report_title in ("Jumplists"):
            workbook_local =  jumplist_report_handler(output_dir_lvariable, workbook_local,report_title)
        elif report_title in ('Shellbags'):
            workbook_local = shellbags_report_handler(output_dir_lvariable, workbook_local,report_title)
        else:
            max_rows_per_sheet = 500000
            try:
                sheet = workbook_local[report_title]
            except:
                sheet = workbook_local.create_sheet(title=report_title)

            additional_sheet_num = 2
            # open csv file and read rows
            with open(output_dir_lvariable+determine_file_to_open(report_title), 'r', encoding=enc) as f:
                reader = csv.reader(f)
                header = next(reader)
                logging.info("Headers = " + str(header))
                sheet.append(header)
                chunk, chunksize, num_rows = [], 50000, 0
                for row in reader:
                    chunk.append(row)
                    if len(chunk)%chunksize==0:
                        logging.info(f"{str(num_rows+chunksize)} rows chunked.")
                        for row in chunk:
                            try:
                                sheet.append(row)
                                num_rows += 1
                            except IllegalCharacterError as e:
                                illegal_char, revised_row = str(e).split(" cannot be used in worksheets.")[0], []
                                for element in row:
                                    if illegal_char in element:
                                        revised_row.append(element.replace(illegal_char, '!'))
                                        logging.warning(f"Removed bad character from row {str(num_rows)}")
                                    else:
                                        revised_row.append(element)
                                sheet.append(revised_row)
                                num_rows += 1
                        if num_rows % max_rows_per_sheet == 0:
                            logging.info(f"Max rows reached! Creating new sheet {report_title}{str(additional_sheet_num)}")
                            sheet = workbook_local.create_sheet(title=f'{report_title}{str(additional_sheet_num)}')
                            sheet.append(header)
                            additional_sheet_num += 1
                        del chunk
                        chunk = []
                        logging.info("Saving workbook...")
                        workbook.save(output_dir_lvariable + 'output.xlsx')
                        logging.info("Workbook saved to " + output_dir_lvariable + 'output.xlsx')
                for row in chunk:
                    try:
                        sheet.append(row)
                    except IllegalCharacterError as e:
                        illegal_char, revised_row = e.args[0].split(" cannot be used in worksheets.")[0], []
                        for element in row:
                            if illegal_char in element:
                                revised_row.append(element.replace(illegal_char, '!'))
                            else:
                                revised_row.append(element)
                        try:
                            sheet.append(revised_row)
                            logging.warning(f"Removed bad character from row {str(num_rows)}")
                        except IllegalCharacterError as e:
                            logging.error(f"Could not append row {str(num_rows)} to sheet due to illegal character.")
                    num_rows += 1
                del chunk
            logging.info(f"{report_title} FINISHED! {str(num_rows)} total rows written to workbook.")
    return workbook_local


def make_local_copies(file_paths,output_dir_local,image_file_drive_letter_local):
    local_outdir = output_dir_local+'templocal'
    try:
        os.mkdir(local_outdir)
    except:
        pass
    for file_path in file_paths:
        logging.info(f"Copying {file_path} to local directory...")
        dest = local_outdir+'\\'+file_path.replace(image_file_drive_letter_local,'')
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        shutil.copy(file_path, dest)
    return local_outdir


def get_shellbag_paths(image_file_drive_letter_local):
    return_set = set()
    for user in os.listdir(image_file_drive_letter_local + 'users'):
        if user not in ('Public', 'All Users', 'Default', 'Default User', 'desktop.ini'):
            return_set.update([image_file_drive_letter_local + 'users\\' + user + '\\NTUSER.dat',
                               image_file_drive_letter_local + 'users\\' + user + '\\NTUSER.dat.log1',
                               image_file_drive_letter_local + 'users\\' + user + '\\NTUSER.dat.log2',
          image_file_drive_letter_local + 'users\\' + user + '\\AppData\\Local\\Microsoft\Windows\\UsrClass.dat',
          image_file_drive_letter_local + 'users\\' + user + '\\AppData\\Local\\Microsoft\Windows\\UsrClass.dat.log1',
            image_file_drive_letter_local + 'users\\' + user + '\\AppData\\Local\\Microsoft\Windows\\UsrClass.dat.log2'])
    return return_set


if __name__ == "__main__":
    report_start_time = get_current_time_as_string()
    logging_configuration(report_start_time)
    IMAGE_FILE, TOOLS_DIR, OUTPUT_DIR = argument_handler()
    if not net6_installed():
        logging.warning(".NET6 RUNTIME NOT INSTALLED, DOWNLOADING AND INSTALLING!")
        install_net6(TOOLS_DIR)
    image_file_drive_letter = mount_image_file_locally(IMAGE_FILE, TOOLS_DIR)
    logging.info("NTFS Image file successfully mounted, beginning tool execution.")
    tools_and_associated_urls = {
    "MFTECmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/MFTECmd.zip",
                    image_file_drive_letter+'$mft', 'MFT'],
    "AmcacheParser.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/AmcacheParser.zip",
                          image_file_drive_letter+'windows\\appcompat\\programs\\amcache.hve', 'Amcache'],
    "AppCompatCacheParser.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/AppCompatCacheParser.zip",
                                image_file_drive_letter+'windows\\system32\\config\\system', 'Shimcache'],
    "EvtxECmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/EvtxECmd.zip",
                     [image_file_drive_letter+'windows\\system32\\winevt\\logs\\'], "EventLogs"],
    "PECmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/PECmd.zip",
                  [image_file_drive_letter+'windows\\prefetch'], "Prefetch"],
    "RBCmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/RBCmd.zip",
                  [image_file_drive_letter+'$recycle.bin'], 'RecycleBin'],
    "SrumECmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/SrumECmd.zip",
                    image_file_drive_letter+'windows\\system32\\sru\\srudb.dat',"SRUM"],
    "BrowsingHistoryView.exe": ["https://www.nirsoft.net/utils/browsinghistoryview.zip",
                                image_file_drive_letter+'users\\',"BrowsingHistory"],
    "JLECmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/JLECmd.zip",
                   set([image_file_drive_letter + 'users\\' + user + '\\appdata\\roaming\\'
                        for user in os.listdir(image_file_drive_letter + 'users')
                        if user not in ('Public', 'All Users', 'Default', 'Default User', 'desktop.ini')]),"Jumplist"],
    "LECmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/LECmd.zip",
                  set([image_file_drive_letter + 'users\\' + user + '\\appdata\\roaming\\'
                       for user in os.listdir(image_file_drive_letter + 'users')
                       if user not in ('Public', 'All Users', 'Default', 'Default User', 'desktop.ini')]),"LNK"],
    "SBECmd.exe": ["https://f001.backblazeb2.com/file/EricZimmermanTools/net6/SBECmd.zip",
                   [make_local_copies(get_shellbag_paths(image_file_drive_letter),OUTPUT_DIR,image_file_drive_letter)], "Shellbags"]}
    workbook = Workbook()
    for tool in tools_and_associated_urls.keys():
        middle_dir = ""
        if tool in ("EvtxECmd.exe"):
            middle_dir = "EvtxECmd\\"
        if not tool_installed(tool, TOOLS_DIR):
            logging.warning(tool + " not installed, downloading...")
            temp_tool_filepath = download_url(tools_and_associated_urls[tool][0], TOOLS_DIR)
            unzip_file(temp_tool_filepath, TOOLS_DIR, delete_original=True)
            logging.info(tool + " downloaded and installed.")
        if isinstance(tools_and_associated_urls[tool][1],set):
            for possible_path in tools_and_associated_urls[tool][1]:
                workbook = execute_tool(TOOLS_DIR+middle_dir+tool,
                                        [possible_path],
                                        OUTPUT_DIR,
                                        workbook,
                                        tools_and_associated_urls[tool][2])
                logging.info("Saving workbook...")
                workbook.save(OUTPUT_DIR + 'output.xlsx')
                logging.info("Workbook saved to "+OUTPUT_DIR + 'output.xlsx')
        else:
            workbook = execute_tool(TOOLS_DIR+middle_dir+tool,
                                    tools_and_associated_urls[tool][1],
                                    OUTPUT_DIR,
                                    workbook,
                                    tools_and_associated_urls[tool][2])
            logging.info("Saving workbook...")
            workbook.save(OUTPUT_DIR + 'output.xlsx')
            logging.info("Workbook saved to " + OUTPUT_DIR + 'output.xlsx')